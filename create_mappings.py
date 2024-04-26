""" Copyright (c) 2020 Cisco and/or its affiliates.
This software is licensed to you under the terms of the Cisco Sample
Code License, Version 1.1 (the "License"). You may obtain a copy of the
License at
           https://developer.cisco.com/docs/licenses
All use of the material herein must be in accordance with the terms of
the License. All rights not expressly granted by the License are
reserved. Unless required by applicable law or agreed to separately in
writing, software distributed under the License is distributed on an "AS
IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express
or implied.
"""
import openpyxl, json,os
from dotenv import load_dotenv
load_dotenv()
MAPPING_FILE = os.environ['MAPPINGS_PATH']

dataframe = openpyxl.load_workbook(MAPPING_FILE)

excel_sheets=[]
templates_list={}
for sheet in dataframe.worksheets:
    excel_sheets.append(sheet.title)
    templates_list[sheet.title]=[]

for sheet in excel_sheets:
    dataframe1 = dataframe[sheet]
    templates_list[sheet]=[]
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        entry=()
        if str(col[1].value) != "None":
            entry=(str(col[1].value),str(col[0].value))
            templates_list[sheet].append(entry)

# print(json.dumps(templates_list, indent=2))